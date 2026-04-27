from fastapi import APIRouter, Depends, UploadFile, File, Query
from sqlalchemy.orm import Session
from typing import Optional
import math, io
from openpyxl import load_workbook
from app.db.session import get_db
from app.dependencies import get_current_user, require_admin
from app.models.user import User
from app.schemas.employee import EmployeeCreate, EmployeeUpdate, EmployeeResponse, EmployeeDetailResponse, BulkImportResult
from app.schemas.common import PaginatedResponse
from app.services import employee_service

router = APIRouter(prefix="/employees", tags=["employees"])


@router.get("", response_model=PaginatedResponse[EmployeeResponse])
def list_employees(
    page: int = Query(1, ge=1),
    limit: int = Query(25, ge=1, le=100),
    search: str = Query(""),
    status: str = Query(""),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    items, total = employee_service.list_employees(db, page, limit, search, status)
    return PaginatedResponse(items=items, total=total, page=page, limit=limit, pages=math.ceil(total / limit))


@router.post("", response_model=EmployeeDetailResponse, status_code=201)
def create_employee(data: EmployeeCreate, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    return employee_service.create_employee(db, data)


@router.get("/{employee_id}", response_model=EmployeeDetailResponse)
def get_employee(employee_id: str, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    return employee_service.get_employee(db, employee_id)


@router.put("/{employee_id}", response_model=EmployeeDetailResponse)
def update_employee(employee_id: str, data: EmployeeUpdate, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    return employee_service.update_employee(db, employee_id, data)


@router.delete("/{employee_id}", status_code=204)
def delete_employee(employee_id: str, db: Session = Depends(get_db), _: User = Depends(require_admin)):
    employee_service.delete_employee(db, employee_id)


@router.post("/bulk-import", response_model=BulkImportResult)
def bulk_import(file: UploadFile = File(...), db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    content = file.file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    headers = [str(c.value).strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            rows.append(dict(zip(headers, [str(v) if v is not None else None for v in row])))
    return employee_service.bulk_import_employees(db, rows)
