from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO


def _header_style(ws, headers: list, row: int = 1):
    fill = PatternFill("solid", fgColor="1e40af")
    font = Font(color="FFFFFF", bold=True)
    border = Border(
        bottom=Side(style="thin"),
        right=Side(style="thin"),
    )
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border


def export_employees(employees: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    headers = ["Employee ID", "First Name", "Last Name", "Email", "Phone",
               "Designation", "Department", "Joining Date", "Status"]
    _header_style(ws, headers)
    for emp in employees:
        ws.append([
            emp.get("employee_id"), emp.get("first_name"), emp.get("last_name"),
            emp.get("email"), emp.get("phone"), emp.get("designation"),
            emp.get("department"), str(emp.get("joining_date", "")), emp.get("status"),
        ])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_attendance(records: list, month_year: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = f"Attendance {month_year}"
    headers = ["Employee ID", "Name", "Date", "Status", "Remarks", "Approval Status"]
    _header_style(ws, headers)
    for r in records:
        ws.append([
            r.get("employee_id"), r.get("employee_name"), str(r.get("attendance_date")),
            r.get("status"), r.get("remarks"), r.get("approval_status"),
        ])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_leave_balance(balances: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Leave Balances"
    headers = ["Employee ID", "Employee Name", "Leave Type", "Available", "Utilized", "Total"]
    _header_style(ws, headers)
    for b in balances:
        total = float(b.get("available_balance", 0)) + float(b.get("utilized_balance", 0))
        ws.append([
            b.get("employee_id"), b.get("employee_name"), b.get("leave_type"),
            float(b.get("available_balance", 0)), float(b.get("utilized_balance", 0)), total,
        ])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_payroll(details: list, month_year: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = f"Payroll {month_year}"
    headers = ["Employee ID", "Name", "Working Days", "Basic", "Allowances",
               "Gross", "PF", "ESIC", "TDS", "Net Salary"]
    _header_style(ws, headers)
    for d in details:
        ws.append([
            d.get("employee_id"), d.get("employee_name"), d.get("working_days"),
            float(d.get("basic_salary", 0)), float(d.get("allowances", 0)),
            float(d.get("gross_salary", 0)), float(d.get("pf_deduction", 0)),
            float(d.get("esic_deduction", 0)), float(d.get("tds_deduction", 0)),
            float(d.get("net_salary", 0)),
        ])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
